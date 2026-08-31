# Tests will be added alongside model implementations.
import pytest
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.assets.models import MaterialType
from apps.inventory.models import InventoryItem


@pytest.fixture
def ops(db):
    return User.objects.create_user(username="inv-ops", password="x", role="ops_manager")


def _client(user):
    c = APIClient()
    c.force_authenticate(user)
    return c


@pytest.fixture
def items(db):
    cable = MaterialType.objects.create(name="Inv Cable", unit="meter")
    mount = MaterialType.objects.create(name="Inv Mount", unit="piece")
    screw = MaterialType.objects.create(name="Inv Screw", unit="box")
    return [
        InventoryItem.objects.create(material_type=cable, quantity=10, unit_cost=100),  # value 1000
        InventoryItem.objects.create(material_type=mount, quantity=50, unit_cost=5),    # value 250
        InventoryItem.objects.create(material_type=screw, quantity=999, unit_cost=None),  # unpriced
    ]


@pytest.mark.django_db
def test_items_expose_per_row_total_value(ops, items):
    r = _client(ops).get("/api/inventory/items/", {"page_size": 100})
    assert r.status_code == 200, r.content
    by_name = {row["material_name"]: row for row in r.data["results"]}
    assert float(by_name["Inv Cable"]["total_value"]) == 1000
    assert float(by_name["Inv Mount"]["total_value"]) == 250
    assert by_name["Inv Screw"]["total_value"] is None


@pytest.mark.django_db
def test_items_order_by_total_value_desc_puts_unpriced_last(ops, items):
    r = _client(ops).get("/api/inventory/items/", {"ordering": "-total_value", "page_size": 100})
    assert r.status_code == 200, r.content
    names = [row["material_name"] for row in r.data["results"]]
    assert names == ["Inv Cable", "Inv Mount", "Inv Screw"]


@pytest.mark.django_db
def test_items_order_by_total_value_asc(ops, items):
    r = _client(ops).get("/api/inventory/items/", {"ordering": "total_value", "page_size": 100})
    names = [row["material_name"] for row in r.data["results"]]
    assert names == ["Inv Screw", "Inv Mount", "Inv Cable"]


# ── Wave 2: Issuance project fields + legacy flow regression ─────────

@pytest.mark.django_db
def test_legacy_direct_issuance_still_decrements_and_journals(ops, items):
    from apps.inventory.models import Issuance, StockMovement

    cable = items[0]  # quantity 10
    c = _client(ops)
    r = c.post("/api/inventory/issuances/", {
        "item": str(cable.pk), "quantity": 4, "reason": "Site consumption",
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["issued_to_project"] is None
    assert r.data["bom_line"] is None
    assert r.data["project_name"] is None

    cable.refresh_from_db()
    assert cable.quantity == 6
    issuance = Issuance.objects.get(pk=r.data["id"])
    movement = StockMovement.objects.get(item=cable, movement_type="out")
    assert movement.quantity == 4
    assert movement.reference == issuance.issue_number

    # over-issue is still rejected
    r = c.post("/api/inventory/issuances/", {
        "item": str(cable.pk), "quantity": 999,
    }, format="json")
    assert r.status_code == 400


@pytest.mark.django_db
def test_direct_issuance_accepts_project_and_bom_line(ops, items):
    from apps.teams.models import Project, ProjectBOMLine

    project = Project.objects.create(name="Issuance Project")
    line = ProjectBOMLine.objects.create(project=project, description="Cable", quantity=3)
    cable = items[0]
    c = _client(ops)
    r = c.post("/api/inventory/issuances/", {
        "item": str(cable.pk), "quantity": 2,
        "issued_to_project": str(project.pk), "bom_line": str(line.pk),
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["issued_to_project"] == project.pk
    assert r.data["bom_line"] == line.pk
    assert r.data["project_name"] == "Issuance Project"
    cable.refresh_from_db()
    assert cable.quantity == 8


# ── Wave 2: goods receipt model gains PO linkage; legacy flow must survive ──

@pytest.mark.django_db
def test_legacy_goods_receipt_still_requires_item_and_increments_stock(ops, items):
    from apps.inventory.models import GoodsReceipt, StockMovement

    cable = items[0]  # quantity 10
    c = _client(ops)
    r = c.post("/api/inventory/receipts/", {
        "item": str(cable.pk), "quantity": 5, "reference": "Legacy DN",
    }, format="json")
    assert r.status_code == 201, r.content
    body = r.data
    assert body["grn_number"]
    assert body["purchase_order"] is None
    assert body["lines"] == []

    cable.refresh_from_db()
    assert cable.quantity == 15
    receipt = GoodsReceipt.objects.get(pk=body["id"])
    assert receipt.received_by == ops
    movement = StockMovement.objects.get(item=cable, movement_type="in")
    assert movement.quantity == 5
    assert movement.reference == receipt.grn_number

    # item/quantity stay mandatory on the legacy endpoint even though the
    # model now allows null for PO-level receipts
    assert c.post("/api/inventory/receipts/", {"quantity": 5}, format="json").status_code == 400
    assert c.post("/api/inventory/receipts/", {"item": str(cable.pk)}, format="json").status_code == 400


# ── Excel export (XC-01) ──────────────────────────────────────────────

import io as _io

from openpyxl import load_workbook as _load_workbook

from apps.accounts.models import AuditLog as _AuditLog


def _sheet_rows(resp):
    wb = _load_workbook(_io.BytesIO(resp.content), read_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


@pytest.mark.django_db
def test_inventory_export_happy_path(ops, items):
    resp = _client(ops).get("/api/inventory/items/export/")
    assert resp.status_code == 200, resp.content
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    rows = _sheet_rows(resp)
    assert rows[0] == [
        "SKU", "Material", "Category", "Quantity", "Min Stock Level",
        "Location", "Unit Cost", "Low Stock",
    ]
    assert len(rows) == 4  # header + 3 items
    by_material = {r[1]: r for r in rows[1:]}
    assert by_material["Inv Cable"][3] == 10
    assert by_material["Inv Cable"][7] == "No"

    log = _AuditLog.objects.filter(action="export", resource_type="inventory_item").latest("created_at")
    assert log.detail["count"] == 3
    assert log.user_id == ops.id


@pytest.mark.django_db
def test_inventory_export_applies_filters_and_low_stock_flag(ops, items):
    transit_type = MaterialType.objects.create(name="Inv Transit Panel", unit="piece")
    InventoryItem.objects.create(
        material_type=transit_type, quantity=2, min_stock_level=5, location="in_transit"
    )

    resp = _client(ops).get("/api/inventory/items/export/", {"location": "in_transit"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert len(rows) == 2
    assert rows[1][1] == "Inv Transit Panel"
    assert rows[1][5] == "In Transit"
    assert rows[1][7] == "Yes"  # 2 <= min_stock_level 5

    log = _AuditLog.objects.filter(action="export", resource_type="inventory_item").latest("created_at")
    assert log.detail == {"count": 1, "params": {"location": "in_transit"}}


# ── Wave 5: ?low_stock= drill-down + category search parity ──────────


@pytest.fixture
def stocked_items(db):
    low_type = MaterialType.objects.create(name="Inv Low Panel", unit="piece")
    ok_type = MaterialType.objects.create(name="Inv OK Panel", unit="piece")
    low = InventoryItem.objects.create(material_type=low_type, quantity=3, min_stock_level=5)
    ok = InventoryItem.objects.create(material_type=ok_type, quantity=50, min_stock_level=5)
    return low, ok


@pytest.mark.django_db
def test_items_low_stock_filter(ops, stocked_items):
    c = _client(ops)
    r = c.get("/api/inventory/items/", {"low_stock": "true", "page_size": 100})
    assert r.status_code == 200, r.content
    names = {row["material_name"] for row in r.data["results"]}
    assert names == {"Inv Low Panel"}

    r = c.get("/api/inventory/items/", {"low_stock": "false", "page_size": 100})
    names = {row["material_name"] for row in r.data["results"]}
    assert names == {"Inv OK Panel"}

    # Unknown values are ignored — both rows come back.
    r = c.get("/api/inventory/items/", {"low_stock": "maybe", "page_size": 100})
    assert len(r.data["results"]) == 2


@pytest.mark.django_db
def test_items_low_stock_boundary_is_inclusive(ops, db):
    edge_type = MaterialType.objects.create(name="Inv Edge Panel", unit="piece")
    InventoryItem.objects.create(material_type=edge_type, quantity=5, min_stock_level=5)

    r = _client(ops).get("/api/inventory/items/", {"low_stock": "true", "page_size": 100})
    assert "Inv Edge Panel" in {row["material_name"] for row in r.data["results"]}


@pytest.mark.django_db
def test_items_low_stock_applies_to_export(ops, stocked_items):
    resp = _client(ops).get("/api/inventory/items/export/", {"low_stock": "true"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert len(rows) == 2  # header + the one low item
    assert rows[1][1] == "Inv Low Panel"
    assert rows[1][7] == "Yes"

    log = _AuditLog.objects.filter(action="export", resource_type="inventory_item").latest("created_at")
    assert log.detail == {"count": 1, "params": {"low_stock": "true"}}


@pytest.mark.django_db
def test_items_search_by_category_name(ops, db):
    from apps.inventory.models import InventoryCategory

    cat = InventoryCategory.objects.create(name="Fasteners")
    bolt = MaterialType.objects.create(name="Inv Bolt", unit="box")
    InventoryItem.objects.create(material_type=bolt, category=cat, quantity=10)
    other = MaterialType.objects.create(name="Inv Other", unit="box")
    InventoryItem.objects.create(material_type=other, quantity=10)

    r = _client(ops).get("/api/inventory/items/", {"search": "Fasteners", "page_size": 100})
    assert r.status_code == 200, r.content
    names = {row["material_name"] for row in r.data["results"]}
    assert names == {"Inv Bolt"}
