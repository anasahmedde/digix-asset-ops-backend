# Tests will be added alongside feature development.
from datetime import timedelta

import pytest
from django.utils import timezone
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.assets.models import Brand, Device, DeviceModel
from apps.clients.models import Client
from apps.sites.models import DeviceInstallation, InstallationStep, Site
from apps.sites.tasks import escalate_overdue_installations


@pytest.fixture
def ops(db):
    return User.objects.create_user(username="site-ops", password="x", role="ops_manager")


@pytest.fixture
def tech(db):
    return User.objects.create_user(
        username="site-tech", password="x", role="technician", first_name="Tariq", last_name="Installer"
    )


@pytest.fixture
def installation(db, tech):
    site = Site.objects.create(name="Install Site", city="Karachi")
    brand = Brand.objects.create(name="InstBrand")
    dm = DeviceModel.objects.create(brand=brand, name="I-1")
    primary = Client.objects.create(name="Primary Client", contact_person="Ali POC", contact_phone="0300-1234567")
    extra = Client.objects.create(name="Second Client")
    device = Device.objects.create(
        device_model=dm, asset_code="AST-INST-1", serial_number="INST-1",
        display_name="Mall Entrance Screen", assigned_client=primary,
    )
    device.clients.add(extra)
    return DeviceInstallation.objects.create(
        device=device, site=site, installed_by=tech, installed_at=timezone.now()
    )


def _client(user):
    c = APIClient()
    c.force_authenticate(user)
    return c


@pytest.mark.django_db
def test_list_exposes_tracker_columns(ops, installation):
    r = _client(ops).get("/api/sites/installations/")
    assert r.status_code == 200, r.content
    row = r.data["results"][0]
    assert row["asset_name"] == "Mall Entrance Screen"
    assert row["client_names"] == ["Primary Client", "Second Client"]
    assert row["poc_name"] == "Ali POC"
    assert row["poc_phone"] == "0300-1234567"
    assert row["installed_by_name"] == "Tariq Installer"
    assert row["due_date"] is None
    assert row["completed_at"] is None
    assert row["client_delays"] == 0


@pytest.mark.django_db
def test_search_by_client_and_installer(ops, installation):
    c = _client(ops)
    assert c.get("/api/sites/installations/", {"search": "Second Client"}).data["count"] == 1
    assert c.get("/api/sites/installations/", {"search": "Tariq"}).data["count"] == 1
    assert c.get("/api/sites/installations/", {"search": "Mall Entrance"}).data["count"] == 1
    assert c.get("/api/sites/installations/", {"search": "no-such-thing"}).data["count"] == 0


@pytest.mark.django_db
def test_due_date_writable_by_manager(ops, installation):
    r = _client(ops).patch(
        f"/api/sites/installations/{installation.id}/", {"due_date": "2026-09-15"}, format="json"
    )
    assert r.status_code == 200, r.content
    installation.refresh_from_db()
    assert str(installation.due_date) == "2026-09-15"


@pytest.mark.django_db
def test_completed_at_stamped_and_cleared_with_steps(installation):
    steps = list(installation.steps.order_by("step_number"))
    assert len(steps) == 6
    for step in steps[:-1]:
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    installation.refresh_from_db()
    assert installation.completed_at is None

    steps[-1].status = InstallationStep.StepStatus.SKIPPED
    steps[-1].save()
    installation.refresh_from_db()
    assert installation.completed_at is not None

    # Reopening a step clears the completion stamp again.
    steps[0].status = InstallationStep.StepStatus.IN_PROGRESS
    steps[0].save()
    installation.refresh_from_db()
    assert installation.completed_at is None


@pytest.mark.django_db
def test_technician_can_log_client_delay(tech, installation):
    step = installation.steps.first()
    r = _client(tech).post("/api/sites/installation-delays/", {
        "installation": str(installation.id),
        "step": str(step.id),
        "cause": "client",
        "description": "Client did not grant site access.",
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["reported_by_name"] == "Tariq Installer"
    assert r.data["cause_display"] == "Client"

    listing = _client(tech).get("/api/sites/installations/")
    assert listing.data["results"][0]["client_delays"] == 1


@pytest.mark.django_db
def test_delay_step_must_belong_to_installation(tech, installation):
    other_site = Site.objects.create(name="Other Site", city="Lahore")
    other = DeviceInstallation.objects.create(
        device=installation.device, site=other_site, installed_at=timezone.now()
    )
    foreign_step = other.steps.first()
    r = _client(tech).post("/api/sites/installation-delays/", {
        "installation": str(installation.id),
        "step": str(foreign_step.id),
        "cause": "client",
    }, format="json")
    assert r.status_code == 400


@pytest.mark.django_db
def test_installer_notified_on_assignment(tech, installation):
    from apps.notifications.models import Notification

    # Created with installed_by=tech (fixture) → one assignment notification.
    notifs = Notification.objects.filter(
        recipient=tech, notification_type="installation_assigned"
    )
    assert notifs.count() == 1
    assert installation.device.asset_code in notifs.first().message

    # Unrelated save must not re-notify.
    installation.notes = "touched"
    installation.save()
    assert notifs.count() == 1

    # Reassignment notifies the new installer.
    other = User.objects.create_user(username="site-tech-2", password="x", role="technician")
    installation.installed_by = other
    installation.save()
    assert Notification.objects.filter(
        recipient=other, notification_type="installation_assigned"
    ).count() == 1


@pytest.mark.django_db
def test_installer_phone_exposed(ops, tech, installation):
    tech.phone = "0301-7654321"
    tech.save()
    r = _client(ops).get("/api/sites/installations/")
    assert r.data["results"][0]["installed_by_phone"] == "0301-7654321"
    r = _client(ops).get(f"/api/sites/installations/{installation.id}/")
    assert r.data["installed_by_phone"] == "0301-7654321"


@pytest.mark.django_db
def test_step_update_restricted_to_installer_or_super_admin(ops, tech, installation):
    step = installation.steps.first()
    # ops manager may NOT advance steps from desktop
    r = _client(ops).patch(f"/api/sites/installation-steps/{step.id}/", {"status": "in_progress"}, format="json")
    assert r.status_code == 403
    # assigned installer may
    r = _client(tech).patch(f"/api/sites/installation-steps/{step.id}/", {"status": "in_progress"}, format="json")
    assert r.status_code == 200, r.content
    # super admin may (incl. the new on-hold status)
    admin = User.objects.create_user(username="site-admin", password="x", role="super_admin")
    r = _client(admin).patch(f"/api/sites/installation-steps/{step.id}/", {"status": "on_hold"}, format="json")
    assert r.status_code == 200, r.content
    step.refresh_from_db()
    assert step.status == "on_hold"


@pytest.mark.django_db
def test_delay_create_restricted(ops, installation):
    r = _client(ops).post("/api/sites/installation-delays/", {
        "installation": str(installation.id), "cause": "client",
    }, format="json")
    assert r.status_code == 403


@pytest.mark.django_db
def test_custom_step_pipeline(ops, installation):
    r = _client(ops).post("/api/sites/installations/", {
        "device": str(installation.device_id),
        "site": str(installation.site_id),
        "installed_at": timezone.now().isoformat(),
        "step_types": ["survey", "programming", "handover"],
    }, format="json")
    assert r.status_code == 201, r.content
    steps = r.data["steps"]
    assert [s["step_type"] for s in steps] == ["survey", "programming", "handover"]
    assert [s["step_number"] for s in steps] == [1, 2, 3]


@pytest.mark.django_db
def test_custom_named_steps_and_vendor(ops, installation):
    from apps.suppliers.models import Supplier

    vendor = Supplier.objects.create(name="Rigging Co")
    r = _client(ops).post("/api/sites/installations/", {
        "device": str(installation.device_id),
        "site": str(installation.site_id),
        "installed_at": timezone.now().isoformat(),
        "vendor": str(vendor.pk),
        "step_types": ["survey", "Crane lift", "handover"],
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["vendor_name"] == "Rigging Co"
    steps = r.data["steps"]
    assert [s["step_type"] for s in steps] == ["survey", "other", "handover"]
    assert steps[1]["step_type_display"] == "Crane lift"


@pytest.mark.django_db
def test_on_hold_steps_count_in_list(ops, installation):
    step = installation.steps.first()
    step.status = InstallationStep.StepStatus.ON_HOLD
    step.save()
    r = _client(ops).get("/api/sites/installations/")
    row = next(x for x in r.data["results"] if x["id"] == str(installation.id))
    assert row["on_hold_steps"] == 1


@pytest.mark.django_db
def test_installation_creation_flips_device_to_installed(installation):
    device = installation.device
    device.refresh_from_db()
    # fixture device starts as procured; creating the installation puts it
    # on the installation track and the flip is journalled
    assert device.status == "installed"
    event = device.lifecycle_events.get(event_type="status_change", to_value="installed")
    assert event.from_value == "procured"
    assert "Install Site" in event.description


@pytest.mark.django_db
def test_completion_marks_device_installed(installation):
    device = installation.device
    device.status = "in_stock"
    device.save()
    for step in installation.steps.all():
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    device.refresh_from_db()
    assert device.status == "installed"


def _complete_non_handover_steps(installation):
    for step in installation.steps.exclude(step_type=InstallationStep.StepType.HANDOVER):
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()


def test_handover_happy_path(installation, ops):
    _complete_non_handover_steps(installation)
    client = _client(ops)
    resp = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Mr. Client POC", "acceptance_notes": "All good"},
        format="multipart",
    )
    assert resp.status_code == 201, resp.data
    body = resp.json()
    assert body["handover"]["accepted_by_name"] == "Mr. Client POC"

    device = installation.device
    device.refresh_from_db()
    installation.refresh_from_db()
    assert device.status == "active"
    assert device.assigned_client_id == installation.handover.client_id
    assert device.current_site_id == installation.site_id
    assert device.installation_date == installation.handover.handover_date
    assert installation.completed_at is not None
    assert installation.steps.get(step_type="handover").status == "completed"
    # Journalled through the machine with the acceptance reason.
    event = device.lifecycle_events.get(event_type="status_change", to_value="active")
    assert "Mr. Client POC" in event.description


def test_handover_blocked_while_steps_pending(installation, ops):
    client = _client(ops)
    resp = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Early Bird"},
        format="multipart",
    )
    assert resp.status_code == 400
    assert "remaining steps" in resp.json()["detail"]


def test_handover_twice_rejected(installation, ops):
    _complete_non_handover_steps(installation)
    client = _client(ops)
    first = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Once"},
        format="multipart",
    )
    assert first.status_code == 201
    again = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Twice"},
        format="multipart",
    )
    assert again.status_code == 400
    assert "already" in again.json()["detail"]


def test_handover_forbidden_for_unassigned_technician(installation):
    _complete_non_handover_steps(installation)
    stranger = User.objects.create_user(username="site-tech-x", password="x", role="technician")
    client = _client(stranger)
    resp = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Nope"},
        format="multipart",
    )
    assert resp.status_code == 403


def test_handover_requires_client_when_device_has_none(installation, ops):
    _complete_non_handover_steps(installation)
    device = installation.device
    device.assigned_client = None
    device.save(update_fields=["assigned_client"])
    client = _client(ops)
    resp = client.post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "No Client"},
        format="multipart",
    )
    assert resp.status_code == 400
    assert "client" in resp.json()


def test_assigned_installer_and_supervisor_can_handover(installation, tech):
    _complete_non_handover_steps(installation)
    r = _client(tech).post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Installer Handover"},
        format="multipart",
    )
    assert r.status_code == 201, r.content

    # A supervisor on a fresh installation works too.
    site2 = Site.objects.create(name="Second Site", city="Lahore")
    device2 = Device.objects.create(
        device_model=installation.device.device_model,
        asset_code="AST-INST-2", serial_number="INST-2",
        assigned_client=installation.device.assigned_client,
    )
    inst2 = DeviceInstallation.objects.create(
        device=device2, site=site2, installed_at=timezone.now()
    )
    _complete_non_handover_steps(inst2)
    supervisor = User.objects.create_user(username="site-super", password="x", role="supervisor")
    r2 = _client(supervisor).post(
        f"/api/sites/installations/{inst2.pk}/handover/",
        {"accepted_by_name": "Supervisor Handover"},
        format="multipart",
    )
    assert r2.status_code == 201, r2.content


def test_handover_reanchors_warranty_even_when_steps_already_done(installation, ops):
    from datetime import timedelta as td

    from apps.warranties.models import Warranty

    today = timezone.localdate()
    warranty = Warranty.objects.create(
        device=installation.device, warranty_type="client", status="active",
        start_date=today, end_date=today + td(days=365), months=12,
    )
    # Close out the WHOLE checklist first (mobile flow), incl. handover step.
    for step in installation.steps.all():
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    installation.refresh_from_db()
    assert installation.completed_at is not None

    paper_date = (today - td(days=30)).isoformat()
    r = _client(ops).post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Paper Acceptance", "handover_date": paper_date},
        format="multipart",
    )
    assert r.status_code == 201, r.content
    warranty.refresh_from_db()
    assert str(warranty.start_date) == paper_date


# ── Wave 4: installation due-date escalation (ES-05) ──────────────────
#
# Policies come from the setup seed migration: scope=installation /
# trigger=due_date — stage 1 (hours=0) -> ops_manager, stage 2 (hours=24)
# -> group_head + ops_manager. Anchor = local midnight ending the due date.


@pytest.fixture
def esc_users(db):
    return {
        "gh": User.objects.create_user(username="inst-esc-gh", password="x", role="group_head"),
        "admin": User.objects.create_user(username="inst-esc-admin", password="x", role="super_admin"),
    }


def _overdue(installation, days):
    DeviceInstallation.objects.filter(pk=installation.pk).update(
        due_date=timezone.localdate() - timedelta(days=days)
    )


def _esc_notifs(installation):
    from apps.notifications.models import Notification

    return Notification.objects.filter(
        installation=installation, notification_type="installation_escalated"
    )


@pytest.mark.django_db
def test_overdue_installation_fires_stage1(ops, tech, installation, esc_users):
    _overdue(installation, days=1)  # anchor = today 00:00 → stage 1 only

    assert escalate_overdue_installations() == 1
    installation.refresh_from_db()
    assert set(installation.escalation_state) == {"due_date:1"}

    notifs = _esc_notifs(installation)
    assert notifs.filter(recipient=tech).exists()  # assigned installer
    assert notifs.filter(recipient=ops).exists()  # escalate_to_role (stage 1)
    assert notifs.filter(recipient=esc_users["admin"]).exists()  # super admin always
    assert not notifs.filter(recipient=esc_users["gh"]).exists()  # group head = stage 2
    assert all(n.is_actionable for n in notifs)


@pytest.mark.django_db
def test_stage2_fires_24h_later_to_group_head(ops, tech, installation, esc_users):
    _overdue(installation, days=2)  # anchor = yesterday 00:00 → both stages elapsed

    assert escalate_overdue_installations() == 2
    installation.refresh_from_db()
    assert set(installation.escalation_state) == {"due_date:1", "due_date:2"}
    assert _esc_notifs(installation).filter(recipient=esc_users["gh"]).exists()


@pytest.mark.django_db
def test_completed_installations_skipped(installation, esc_users):
    DeviceInstallation.objects.filter(pk=installation.pk).update(
        due_date=timezone.localdate() - timedelta(days=5),
        completed_at=timezone.now(),
    )
    assert escalate_overdue_installations() == 0
    installation.refresh_from_db()
    assert installation.escalation_state == {}
    assert not _esc_notifs(installation).exists()


@pytest.mark.django_db
def test_escalation_rerun_is_idempotent(installation, esc_users):
    _overdue(installation, days=3)
    assert escalate_overdue_installations() == 2
    installation.refresh_from_db()
    first_state = dict(installation.escalation_state)
    first_count = _esc_notifs(installation).count()

    assert escalate_overdue_installations() == 0
    installation.refresh_from_db()
    assert installation.escalation_state == first_state  # keys and timestamps untouched
    assert _esc_notifs(installation).count() == first_count


@pytest.mark.django_db
def test_escalated_filter_and_serializer_flag(ops, installation, esc_users):
    _overdue(installation, days=1)
    other = DeviceInstallation.objects.create(
        device=installation.device,
        site=installation.site,
        installed_at=timezone.now(),
    )
    escalate_overdue_installations()

    c = _client(ops)
    hot = c.get("/api/sites/installations/", {"escalated": "true"})
    assert [row["id"] for row in hot.data["results"]] == [str(installation.id)]
    assert hot.data["results"][0]["escalated"] is True
    assert "due_date:1" in hot.data["results"][0]["escalation_state"]

    cold = c.get("/api/sites/installations/", {"escalated": "false"})
    cold_ids = [row["id"] for row in cold.data["results"]]
    assert str(other.id) in cold_ids and str(installation.id) not in cold_ids
    row = next(r for r in cold.data["results"] if r["id"] == str(other.id))
    assert row["escalated"] is False

    detail = c.get(f"/api/sites/installations/{installation.id}/")
    assert detail.data["escalated"] is True
    assert "due_date:1" in detail.data["escalation_state"]


# ── Excel export (XC-01) ──────────────────────────────────────────────

import io as _io

from openpyxl import load_workbook as _load_workbook

from apps.accounts.models import AuditLog as _AuditLog


def _sheet_rows(resp):
    wb = _load_workbook(_io.BytesIO(resp.content), read_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


@pytest.mark.django_db
def test_installations_export_happy_path(ops, installation):
    # New installations auto-seed the 6-step pipeline; complete half of it.
    for step in installation.steps.order_by("step_number")[:3]:
        step.status = "completed"
        step.save()

    resp = _client(ops).get("/api/sites/installations/export/")
    assert resp.status_code == 200, resp.content
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    rows = _sheet_rows(resp)
    assert rows[0][:4] == ["Asset Code", "Asset Name", "Site", "Clients"]
    assert len(rows) == 2
    row = rows[1]
    assert row[0] == "AST-INST-1"
    assert row[1] == "Mall Entrance Screen"
    assert row[2] == "Install Site"
    assert row[3] == "Primary Client, Second Client"
    assert row[4] == "Tariq Installer"
    assert row[8] == 50  # 1 of 2 steps completed
    assert row[9] == "No"  # not escalated

    log = _AuditLog.objects.filter(action="export", resource_type="installation").latest("created_at")
    assert log.detail["count"] == 1
    assert log.user_id == ops.id


@pytest.mark.django_db
def test_installations_export_applies_filters(ops, installation, tech):
    other_site = Site.objects.create(name="Other Site", city="Lahore")
    brand = Brand.objects.create(name="ExpBrand")
    dm = DeviceModel.objects.create(brand=brand, name="E-1")
    other_device = Device.objects.create(
        device_model=dm, asset_code="AST-EXP-2", serial_number="EXP-2"
    )
    DeviceInstallation.objects.create(
        device=other_device, site=other_site, installed_by=tech, installed_at=timezone.now()
    )

    resp = _client(ops).get("/api/sites/installations/export/", {"site": str(other_site.id)})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert [r[0] for r in rows[1:]] == ["AST-EXP-2"]

    log = _AuditLog.objects.filter(action="export", resource_type="installation").latest("created_at")
    assert log.detail == {"count": 1, "params": {"site": str(other_site.id)}}


# ── Wave 5: ?bucket= tracker drill-downs ──────────────────────────────


@pytest.fixture
def bucket_installations(db, tech):
    """One installation per progress bucket (default 6-step checklist each)."""
    from apps.sites.models import InstallationDelay

    site = Site.objects.create(name="Bucket Site", city="Karachi")
    brand = Brand.objects.create(name="BucketBrand")
    dm = DeviceModel.objects.create(brand=brand, name="B-1")

    def mk(name, **kwargs):
        device = Device.objects.create(
            device_model=dm, serial_number=f"BKT-{name}", display_name=f"Bucket {name}",
        )
        return DeviceInstallation.objects.create(
            device=device, site=site, installed_by=tech, installed_at=timezone.now(), **kwargs
        )

    fresh = mk("fresh")

    progress = mk("progress")  # two advanced steps — distinct must dedupe
    for step, status in zip(progress.steps.order_by("step_number"), ("in_progress", "completed")):
        step.status = status
        step.save()

    hold = mk("hold")  # two held steps — distinct must dedupe
    for step in hold.steps.order_by("step_number")[:2]:
        step.status = InstallationStep.StepStatus.ON_HOLD
        step.save()

    done = mk("done")
    for step in done.steps.order_by("step_number"):
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    done.refresh_from_db()
    assert done.completed_at is not None  # signal stamped completion

    late = mk("late", due_date=timezone.localdate() - timedelta(days=1))

    delayed = mk("delayed")  # two client delays — distinct must dedupe
    for i in range(2):
        InstallationDelay.objects.create(
            installation=delayed, cause=InstallationDelay.Cause.CLIENT,
            description=f"Client kept the site closed ({i})", reported_by=tech,
        )
    # A non-client delay alone must NOT put an installation in the bucket.
    InstallationDelay.objects.create(
        installation=fresh, cause=InstallationDelay.Cause.VENDOR, description="Vendor late",
    )

    return {
        "fresh": fresh, "progress": progress, "hold": hold,
        "done": done, "late": late, "delayed": delayed,
    }


@pytest.mark.django_db
def test_installation_bucket_filters(ops, bucket_installations):
    c = _client(ops)

    def names(bucket):
        r = c.get("/api/sites/installations/", {"bucket": bucket, "page_size": 100})
        assert r.status_code == 200, r.content
        return [row["asset_name"] for row in r.data["results"]]

    assert names("completed") == ["Bucket done"]
    assert names("overdue") == ["Bucket late"]
    assert names("on_hold") == ["Bucket hold"]  # deduped despite 2 held steps
    assert names("in_progress") == ["Bucket progress"]  # deduped despite 2 steps
    assert sorted(names("not_started")) == ["Bucket delayed", "Bucket fresh", "Bucket late"]
    assert names("delayed") == ["Bucket delayed"]  # client-cause only, deduped

    # Unknown values are ignored — the whole tracker comes back.
    r = c.get("/api/sites/installations/", {"bucket": "bogus", "page_size": 100})
    assert len(r.data["results"]) == len(bucket_installations)


@pytest.mark.django_db
def test_installation_bucket_applies_to_export(ops, bucket_installations):
    resp = _client(ops).get("/api/sites/installations/export/", {"bucket": "overdue"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert [r[1] for r in rows[1:]] == ["Bucket late"]

    log = _AuditLog.objects.filter(action="export", resource_type="installation").latest("created_at")
    assert log.detail == {"count": 1, "params": {"bucket": "overdue"}}


# ── Vendor access (XC-04) ─────────────────────────────────────────────


@pytest.fixture
def vendor_install(db, installation):
    """A second installation done by a vendor, plus vendor logins."""
    from apps.suppliers.models import Supplier

    supplier = Supplier.objects.create(name="Install Vendor")
    other_supplier = Supplier.objects.create(name="Rival Vendor")
    vendor_user = User.objects.create_user(
        username="site-vendor", password="x", role="vendor", supplier=supplier
    )
    other_vendor_user = User.objects.create_user(
        username="site-vendor-b", password="x", role="vendor", supplier=other_supplier
    )
    unlinked_vendor = User.objects.create_user(username="site-vendor-none", password="x", role="vendor")

    site = Site.objects.create(name="Vendor Site", city="Lahore")
    brand = Brand.objects.create(name="VendBrand")
    dm = DeviceModel.objects.create(brand=brand, name="V-1")
    device = Device.objects.create(
        device_model=dm, asset_code="AST-VEND-1", serial_number="VEND-1",
        display_name="Vendor Screen",
    )
    vendor_installation = DeviceInstallation.objects.create(
        device=device, site=site, vendor=supplier, installed_at=timezone.now()
    )
    return {
        "supplier": supplier,
        "vendor_user": vendor_user,
        "other_vendor_user": other_vendor_user,
        "unlinked_vendor": unlinked_vendor,
        "vendor_installation": vendor_installation,
    }


@pytest.mark.django_db
def test_vendor_sees_only_own_installations(vendor_install, installation):
    r = _client(vendor_install["vendor_user"]).get("/api/sites/installations/")
    assert r.status_code == 200
    assert r.data["count"] == 1
    assert r.data["results"][0]["id"] == str(vendor_install["vendor_installation"].id)
    # Detail of the tech-run installation is invisible to the vendor.
    r = _client(vendor_install["vendor_user"]).get(f"/api/sites/installations/{installation.id}/")
    assert r.status_code == 404
    # Other supplier's vendor and an unlinked vendor login see nothing.
    assert _client(vendor_install["other_vendor_user"]).get("/api/sites/installations/").data["count"] == 0
    assert _client(vendor_install["unlinked_vendor"]).get("/api/sites/installations/").data["count"] == 0


@pytest.mark.django_db
def test_vendor_advances_own_step_403_on_others(vendor_install, installation):
    c = _client(vendor_install["vendor_user"])
    own_step = vendor_install["vendor_installation"].steps.first()
    r = c.patch(f"/api/sites/installation-steps/{own_step.id}/", {"status": "in_progress"}, format="json")
    assert r.status_code == 200, r.content
    own_step.refresh_from_db()
    assert own_step.status == "in_progress"

    other_step = installation.steps.first()
    r = c.patch(f"/api/sites/installation-steps/{other_step.id}/", {"status": "in_progress"}, format="json")
    assert r.status_code == 403
    r = _client(vendor_install["other_vendor_user"]).patch(
        f"/api/sites/installation-steps/{own_step.id}/", {"status": "completed"}, format="json"
    )
    assert r.status_code == 403


@pytest.mark.django_db
def test_vendor_uploads_photos_own_installation_only(vendor_install, installation):
    import io

    from PIL import Image

    def _png():
        buf = io.BytesIO()
        Image.new("RGB", (8, 8), "blue").save(buf, format="PNG")
        buf.seek(0)
        buf.name = "site.png"
        return buf

    c = _client(vendor_install["vendor_user"])
    r = c.post("/api/sites/installation-photos/", {
        "installation": str(vendor_install["vendor_installation"].id),
        "photo_type": "pre_install",
        "image": _png(),
    }, format="multipart")
    assert r.status_code == 201, r.content

    r = c.post("/api/sites/installation-photos/", {
        "installation": str(installation.id),
        "photo_type": "pre_install",
        "image": _png(),
    }, format="multipart")
    assert r.status_code == 403


@pytest.mark.django_db
def test_vendor_cannot_handover(vendor_install):
    inst = vendor_install["vendor_installation"]
    for step in inst.steps.exclude(step_type=InstallationStep.StepType.HANDOVER):
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    r = _client(vendor_install["vendor_user"]).post(
        f"/api/sites/installations/{inst.pk}/handover/",
        {"accepted_by_name": "Client POC"},
        format="multipart",
    )
    assert r.status_code == 403


def test_handover_client_falls_back_to_project_then_site(installation, ops):
    from apps.teams.models import Project

    device = installation.device
    project_client = Client.objects.create(name="Project Buyer")
    project = Project.objects.create(name="Fallback Project", client=project_client)
    device.assigned_client = None
    device.project = project
    device.save(update_fields=["assigned_client", "project"])
    _complete_non_handover_steps(installation)
    r = _client(ops).post(
        f"/api/sites/installations/{installation.pk}/handover/",
        {"accepted_by_name": "Fallback POC"},
        format="multipart",
    )
    assert r.status_code == 201, r.content
    device.refresh_from_db()
    assert device.assigned_client_id == project_client.pk
