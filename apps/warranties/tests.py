# Tests will be added alongside model implementations.
from datetime import timedelta

import pytest
from django.utils import timezone
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.assets.models import Brand, Device, DeviceModel
from apps.warranties.models import Warranty
from apps.warranties.tasks import complete_expired_warranties


@pytest.fixture
def ops(db):
    return User.objects.create_user(username="war-ops", password="x", role="ops_manager")


@pytest.fixture
def device(db):
    brand = Brand.objects.create(name="WarBrand")
    dm = DeviceModel.objects.create(brand=brand, name="W-1")
    return Device.objects.create(device_model=dm, asset_code="AST-WAR-1", serial_number="WAR-1")


def _client(user):
    c = APIClient()
    c.force_authenticate(user)
    return c


@pytest.mark.django_db
def test_client_warranty_created_with_device(ops):
    brand = Brand.objects.create(name="WarBrand2")
    dm = DeviceModel.objects.create(brand=brand, name="W-2")
    c = _client(ops)
    r = c.post("/api/assets/devices/", {
        "serial_number": "WAR-NEW-1",
        "asset_code": "AST-WAR-NEW-1",
        "device_model": str(dm.pk),
        "client_warranty_months": 6,
    }, format="json")
    assert r.status_code == 201, r.content
    w = Warranty.objects.get(device__serial_number="WAR-NEW-1")
    assert w.warranty_type == "client"
    assert w.months == 6
    assert (w.end_date - w.start_date).days >= 180


@pytest.mark.django_db
def test_auto_complete_after_term(device):
    w = Warranty.objects.create(
        device=device, warranty_type="client", status="active", months=3,
        start_date=timezone.now().date() - timedelta(days=120),
        end_date=timezone.now().date() - timedelta(days=1),
    )
    assert complete_expired_warranties() == 1
    w.refresh_from_db()
    assert w.status == "expired"
    assert w.get_status_display() == "Warranty Completed"
    # one-shot
    assert complete_expired_warranties() == 0


@pytest.mark.django_db
def test_reissue_flow(device):
    # Client warranties are marketing/admin-side; ops roles no longer see them.
    admin = User.objects.create_user(username="war-admin", password="x", role="super_admin")
    w = Warranty.objects.create(
        device=device, warranty_type="client", status="expired", months=3,
        start_date=timezone.now().date() - timedelta(days=120),
        end_date=timezone.now().date() - timedelta(days=1),
    )
    c = _client(admin)
    r = c.post(f"/api/warranties/{w.pk}/reissue/", {"months": 12}, format="json")
    assert r.status_code == 201, r.content
    w.refresh_from_db()
    assert w.status == "reissued"
    new = Warranty.objects.get(reissued_from=w)
    assert new.warranty_type == "client" and new.months == 12 and new.status == "active"

    # invalid term rejected
    w2 = Warranty.objects.create(
        device=device, warranty_type="client", status="expired",
        start_date=timezone.now().date() - timedelta(days=60),
        end_date=timezone.now().date() - timedelta(days=2),
    )
    r = c.post(f"/api/warranties/{w2.pk}/reissue/", {"months": 5}, format="json")
    assert r.status_code == 400


@pytest.mark.django_db
def test_device_immutable_on_update(ops, device):
    other_brand = Brand.objects.create(name="WarBrand3")
    other_dm = DeviceModel.objects.create(brand=other_brand, name="W-3")
    other_device = Device.objects.create(
        device_model=other_dm, asset_code="AST-WAR-OTHER", serial_number="WAR-OTHER"
    )
    w = Warranty.objects.create(
        device=device, warranty_type="manufacturer", status="active",
        start_date=timezone.now().date(),
        end_date=timezone.now().date() + timedelta(days=365),
    )
    c = _client(ops)
    r = c.patch(f"/api/warranties/{w.pk}/", {
        "device": str(other_device.pk),
        "notes": "updated",
    }, format="json")
    assert r.status_code == 200, r.content
    w.refresh_from_db()
    assert w.device_id == device.pk
    assert w.notes == "updated"
    # reissued_from cannot be forged through the API either
    r = c.patch(f"/api/warranties/{w.pk}/", {"reissued_from": str(w.pk)}, format="json")
    assert r.status_code == 200
    w.refresh_from_db()
    assert w.reissued_from_id is None


@pytest.mark.django_db
def test_device_name_exposed(ops, device):
    device.display_name = "Lobby Screen"
    device.save()
    w = Warranty.objects.create(
        device=device, warranty_type="manufacturer", status="active",
        start_date=timezone.now().date(),
        end_date=timezone.now().date() + timedelta(days=30),
    )
    r = _client(ops).get(f"/api/warranties/{w.pk}/")
    assert r.data["device_name"] == "Lobby Screen"
    assert r.data["device_code"] == device.asset_code


@pytest.mark.django_db
def test_role_scoped_visibility(device):
    Warranty.objects.create(
        device=device, warranty_type="client", status="active", months=6,
        start_date=timezone.now().date(), end_date=timezone.now().date() + timedelta(days=180),
    )
    Warranty.objects.create(
        device=device, warranty_type="supplier", status="active",
        start_date=timezone.now().date(), end_date=timezone.now().date() + timedelta(days=365),
    )
    def types_for(role):
        u = User.objects.create_user(username=f"war-{role}", password="x", role=role)
        r = _client(u).get("/api/warranties/")
        return sorted({row["warranty_type"] for row in r.data["results"]})

    assert types_for("marketing") == ["client"]
    assert types_for("technician") == ["supplier"]
    assert types_for("super_admin") == ["client", "supplier"]
    assert types_for("group_head") == ["client", "supplier"]


@pytest.mark.django_db
def test_supplier_warranty_anchors_to_purchase_date(ops):
    from datetime import date

    brand = Brand.objects.create(name="WarBrand4")
    dm = DeviceModel.objects.create(brand=brand, name="W-4")
    device = Device.objects.create(
        device_model=dm, asset_code="AST-WAR-4", serial_number="WAR-4",
        purchase_date=date(2026, 7, 1),
    )
    r = _client(ops).post("/api/warranties/", {
        "device": str(device.pk),
        "warranty_type": "supplier",
        "months": 12,
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["start_date"] == "2026-07-01"
    assert r.data["end_date"] == "2027-07-01"


@pytest.mark.django_db
def test_client_warranty_reanchors_on_handover(device):
    from apps.sites.models import DeviceInstallation, InstallationStep, Site

    w = Warranty.objects.create(
        device=device, warranty_type="client", status="active", months=6,
        start_date=timezone.now().date() - timedelta(days=90),
        end_date=timezone.now().date() + timedelta(days=90),
    )
    site = Site.objects.create(name="Handover Site", city="Karachi")
    inst = DeviceInstallation.objects.create(device=device, site=site, installed_at=timezone.now())
    for step in inst.steps.all():
        step.status = InstallationStep.StepStatus.COMPLETED
        step.save()
    inst.refresh_from_db()
    assert inst.completed_at is not None
    w.refresh_from_db()
    assert w.start_date == inst.completed_at.date()
    assert (w.end_date - w.start_date).days >= 180


# ── Excel export (XC-01) ──────────────────────────────────────────────

import io as _io

from openpyxl import load_workbook as _load_workbook

from apps.accounts.models import AuditLog as _AuditLog


def _sheet_rows(resp):
    wb = _load_workbook(_io.BytesIO(resp.content), read_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


@pytest.fixture
def mixed_warranties(db, device):
    today = timezone.now().date()
    client_w = Warranty.objects.create(
        device=device, warranty_type="client", start_date=today,
        end_date=today + timedelta(days=90), months=3, reference_number="REF-CL-1",
    )
    supplier_w = Warranty.objects.create(
        device=device, warranty_type="manufacturer", start_date=today,
        end_date=today + timedelta(days=365), months=12, reference_number="REF-MF-1",
    )
    return client_w, supplier_w


@pytest.mark.django_db
def test_warranties_export_happy_path(mixed_warranties):
    admin = User.objects.create_user(username="war-admin", password="x", role="super_admin")
    resp = _client(admin).get("/api/warranties/export/")
    assert resp.status_code == 200, resp.content
    assert resp["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    rows = _sheet_rows(resp)
    assert rows[0][:4] == ["Asset Code", "Component", "Warranty Type", "Status"]
    assert len(rows) == 3  # header + both warranties (admin sees all)

    log = _AuditLog.objects.filter(action="export", resource_type="warranty").latest("created_at")
    assert log.detail["count"] == 2
    assert log.user_id == admin.id


@pytest.mark.django_db
def test_warranties_export_role_scoped(mixed_warranties):
    # Marketing only sees client warranties; ops only supplier-facing ones.
    marketing = User.objects.create_user(username="war-mkt", password="x", role="marketing")
    resp = _client(marketing).get("/api/warranties/export/")
    rows = _sheet_rows(resp)
    assert [r[8] for r in rows[1:]] == ["REF-CL-1"]

    ops = User.objects.create_user(username="war-ops-x", password="x", role="ops_manager")
    resp = _client(ops).get("/api/warranties/export/")
    rows = _sheet_rows(resp)
    assert [r[8] for r in rows[1:]] == ["REF-MF-1"]


@pytest.mark.django_db
def test_warranties_export_applies_filters(mixed_warranties):
    admin = User.objects.create_user(username="war-admin2", password="x", role="super_admin")
    resp = _client(admin).get("/api/warranties/export/", {"warranty_type": "client"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert len(rows) == 2
    assert rows[1][8] == "REF-CL-1"

    log = _AuditLog.objects.filter(action="export", resource_type="warranty").latest("created_at")
    assert log.detail == {"count": 1, "params": {"warranty_type": "client"}}


# ── Wave 5: ?side= drill-down (list + export) ─────────────────────────


@pytest.fixture
def all_sides(db, device):
    """One warranty per type so ?side= grouping is fully exercised."""
    today = timezone.now().date()
    refs = {}
    for wtype, ref in (
        ("client", "REF-SIDE-CL"),
        ("supplier", "REF-SIDE-SU"),
        ("manufacturer", "REF-SIDE-MF"),
        ("extended", "REF-SIDE-EX"),
    ):
        refs[wtype] = Warranty.objects.create(
            device=device, warranty_type=wtype, start_date=today,
            end_date=today + timedelta(days=365), months=12, reference_number=ref,
        )
    return refs


@pytest.mark.django_db
def test_side_filter_on_list(all_sides):
    admin = User.objects.create_user(username="war-side-admin", password="x", role="super_admin")
    c = _client(admin)

    r = c.get("/api/warranties/", {"side": "client", "page_size": 100})
    assert r.status_code == 200, r.content
    refs = {row["reference_number"] for row in r.data["results"]}
    assert refs == {"REF-SIDE-CL"}

    r = c.get("/api/warranties/", {"side": "supplier", "page_size": 100})
    refs = {row["reference_number"] for row in r.data["results"]}
    assert refs == {"REF-SIDE-SU", "REF-SIDE-MF", "REF-SIDE-EX"}

    # Unknown values are ignored — everything comes back.
    r = c.get("/api/warranties/", {"side": "bogus", "page_size": 100})
    assert len(r.data["results"]) == 4


@pytest.mark.django_db
def test_side_filter_respects_role_scope(all_sides):
    # Marketing is already scoped to client warranties — asking for the
    # supplier side must not widen their view.
    marketing = User.objects.create_user(username="war-side-mkt", password="x", role="marketing")
    r = _client(marketing).get("/api/warranties/", {"side": "supplier", "page_size": 100})
    assert r.status_code == 200, r.content
    assert r.data["results"] == []


@pytest.mark.django_db
def test_side_filter_applies_to_export(all_sides):
    admin = User.objects.create_user(username="war-side-admin2", password="x", role="super_admin")
    resp = _client(admin).get("/api/warranties/export/", {"side": "supplier"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert {r[8] for r in rows[1:]} == {"REF-SIDE-SU", "REF-SIDE-MF", "REF-SIDE-EX"}

    log = _AuditLog.objects.filter(action="export", resource_type="warranty").latest("created_at")
    assert log.detail == {"count": 3, "params": {"side": "supplier"}}
