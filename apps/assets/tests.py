import pytest
from rest_framework.test import APIClient

from apps.accounts.models import User
from apps.assets.models import AssetCode, Brand, Device, DeviceModel


@pytest.fixture
def admin_client(db):
    admin = User.objects.create_user(
        username="labeladmin", password="x", role="super_admin"
    )
    client = APIClient()
    client.force_authenticate(admin)
    return client


@pytest.fixture
def device(db):
    brand = Brand.objects.create(name="LabelBrand")
    model = DeviceModel.objects.create(brand=brand, name="LB-55")
    return Device.objects.create(device_model=model, serial_number="SN-LABEL-1")


@pytest.mark.django_db
def test_label_generates_png_qr(admin_client, device):
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/label/", {"format": "qr"}, format="json"
    )
    assert resp.status_code == 200, resp.content
    assert resp.json()["generated_file"]

    label = AssetCode.objects.get(device=device, format="qr", is_current=True)
    with label.generated_file.open("rb") as fh:
        assert fh.read(8).startswith(b"\x89PNG")


@pytest.mark.django_db
def test_label_regeneration_reuses_current_row(admin_client, device):
    for _ in range(2):
        resp = admin_client.post(
            f"/api/assets/devices/{device.id}/label/", {"format": "qr"}, format="json"
        )
        assert resp.status_code == 200
    assert AssetCode.objects.filter(device=device, format="qr").count() == 1


@pytest.mark.django_db
def test_label_code128_and_invalid_format(admin_client, device):
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/label/", {"format": "code128"}, format="json"
    )
    assert resp.status_code == 200
    assert AssetCode.objects.filter(device=device, format="code128").exists()

    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/label/", {"format": "pdf417"}, format="json"
    )
    assert resp.status_code == 400


@pytest.mark.django_db
def test_label_requires_manager_role(device):
    tech = User.objects.create_user(username="labeltech", password="x", role="technician")
    client = APIClient()
    client.force_authenticate(tech)
    resp = client.post(
        f"/api/assets/devices/{device.id}/label/", {"format": "qr"}, format="json"
    )
    assert resp.status_code == 403


# ── Bulk labels (WF-05): one PDF, one label per page ──────────────────


def _pdf_page_count(content: bytes) -> int:
    # Pillow writes one "/Type /Page" object per page plus one "/Type /Pages" node.
    return content.count(b"/Type /Page") - content.count(b"/Type /Pages")


@pytest.fixture
def device_batch(device):
    dm = device.device_model
    others = [
        Device.objects.create(device_model=dm, serial_number=f"SN-BULK-{i}")
        for i in range(2)
    ]
    return [device, *others]


@pytest.mark.django_db
def test_bulk_labels_returns_multipage_pdf(admin_client, device_batch):
    resp = admin_client.post(
        "/api/assets/devices/labels/",
        {"ids": [str(d.id) for d in device_batch]},
        format="json",
    )
    assert resp.status_code == 200, resp.content
    assert resp["Content-Type"] == "application/pdf"
    assert "labels-qr-" in resp["Content-Disposition"]
    assert resp.content.startswith(b"%PDF")
    assert _pdf_page_count(resp.content) == 3


@pytest.mark.django_db
def test_bulk_labels_code128_and_invalid_format(admin_client, device_batch):
    ids = [str(d.id) for d in device_batch]
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids, "format": "code128"}, format="json"
    )
    assert resp.status_code == 200, resp.content
    assert resp.content.startswith(b"%PDF")
    assert "labels-code128-" in resp["Content-Disposition"]

    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids, "format": "pdf417"}, format="json"
    )
    assert resp.status_code == 400
    assert "format" in resp.json()


@pytest.mark.django_db
def test_bulk_labels_rejects_empty_and_oversized_batches(admin_client, device):
    import uuid as _uuid

    for bad_body in ({}, {"ids": []}, {"ids": "not-a-list"}):
        resp = admin_client.post("/api/assets/devices/labels/", bad_body, format="json")
        assert resp.status_code == 400, bad_body

    too_many = [str(_uuid.uuid4()) for _ in range(201)]
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": too_many}, format="json"
    )
    assert resp.status_code == 400
    assert "At most 200" in str(resp.json()["ids"])


@pytest.mark.django_db
def test_bulk_labels_rejects_bad_uuids_and_unknown_ids(admin_client, device):
    import uuid as _uuid

    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ["not-a-uuid"]}, format="json"
    )
    assert resp.status_code == 400

    unknown = _uuid.uuid4()
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": [str(unknown)]}, format="json"
    )
    assert resp.status_code == 400
    assert str(unknown) in str(resp.json()["ids"])


@pytest.mark.django_db
def test_bulk_labels_400_lists_missing_ids(admin_client, device_batch):
    """A partial batch fails loudly, naming every unresolved id (deduped)."""
    import uuid as _uuid

    missing_a, missing_b = _uuid.uuid4(), _uuid.uuid4()
    ids = [str(d.id) for d in device_batch] + [
        str(missing_a), str(missing_b), str(missing_a),  # duplicate collapses
    ]
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids}, format="json"
    )
    assert resp.status_code == 400
    message = str(resp.json()["ids"])
    assert str(missing_a) in message and str(missing_b) in message
    assert message.count(str(missing_a)) == 1  # deduped before reporting
    # No devices resolved → no ledger rows written for the good ids either.
    assert AssetCode.objects.count() == 0


@pytest.mark.django_db
def test_bulk_labels_accepts_duplicate_known_ids(admin_client, device_batch):
    """Duplicated known ids collapse to one page each instead of a 400."""
    ids = [str(d.id) for d in device_batch] + [str(device_batch[0].id)]
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids}, format="json"
    )
    assert resp.status_code == 200, resp.content
    assert _pdf_page_count(resp.content) == 3


@pytest.mark.django_db
def test_bulk_labels_persist_asset_code_ledger(admin_client, device_batch):
    """Bulk prints leave the same AssetCode trail as single-label prints."""
    ids = [str(d.id) for d in device_batch]
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids, "format": "qr"}, format="json"
    )
    assert resp.status_code == 200, resp.content
    for d in device_batch:
        label = AssetCode.objects.get(device=d, format="qr", is_current=True)
        assert label.label_size == "60x30"
        with label.generated_file.open("rb") as fh:
            assert fh.read(8).startswith(b"\x89PNG")

    # Re-printing reuses the current rows — no duplicate ledger entries.
    resp = admin_client.post(
        "/api/assets/devices/labels/", {"ids": ids, "format": "qr"}, format="json"
    )
    assert resp.status_code == 200
    assert AssetCode.objects.filter(format="qr").count() == len(device_batch)


@pytest.mark.django_db
def test_bulk_labels_requires_manager_role(device):
    tech = User.objects.create_user(username="bulktech", password="x", role="technician")
    client = APIClient()
    client.force_authenticate(tech)
    resp = client.post(
        "/api/assets/devices/labels/", {"ids": [str(device.id)]}, format="json"
    )
    assert resp.status_code == 403


# ── Asset composition (Project -> Asset -> Components) ────────────────

import pytest as _pytest
from rest_framework.test import APIClient as _APIClient

from apps.accounts.models import User as _User


@_pytest.mark.django_db
def test_components_and_project_link():
    from apps.assets.models import Brand, Device, DeviceModel
    from apps.teams.models import Project

    ops = _User.objects.create_user(username="cmp-ops", password="x", role="ops_manager")
    brand = Brand.objects.create(name="CmpBrand")
    dm = DeviceModel.objects.create(brand=brand, name="C-1")
    proj = Project.objects.create(name="Cmp Order")
    dev = Device.objects.create(device_model=dm, asset_code="AST-CMP-1", serial_number="CMP-1", project=proj)

    c = _APIClient()
    c.force_authenticate(ops)

    r = c.post("/api/assets/components/", {
        "device": str(dev.pk), "name": "SMD Cabinet P3.9",
        "component_type": "Cabinet", "quantity": 12,
    }, format="json")
    assert r.status_code == 201, r.content
    r = c.post("/api/assets/components/", {
        "device": str(dev.pk), "name": "Media Player", "quantity": 1,
    }, format="json")
    assert r.status_code == 201

    detail = c.get(f"/api/assets/devices/{dev.pk}/").json()
    assert detail["project_name"] == "Cmp Order"
    assert len(detail["components"]) == 2

    projects = c.get("/api/teams/projects/").json()
    rows = projects.get("results", projects)
    row = next(p for p in rows if p["id"] == str(proj.pk))
    assert row["assets_count"] == 1


@pytest.mark.django_db
def test_component_with_supplier_and_warranty(db):
    from datetime import date

    from apps.suppliers.models import Supplier
    from apps.warranties.models import Warranty

    brand = Brand.objects.create(name="CompBrand")
    dm = DeviceModel.objects.create(brand=brand, name="C-1")
    supplier = Supplier.objects.create(name="Comp Supplier")
    device = Device.objects.create(
        device_model=dm, asset_code="AST-COMP-1", serial_number="COMP-1",
        purchase_date=date(2026, 8, 1),
    )
    ops = User.objects.create_user(username="comp-ops", password="x", role="ops_manager")
    client = APIClient()
    client.force_authenticate(ops)
    r = client.post("/api/assets/components/", {
        "device": str(device.pk),
        "name": "Receiving Card",
        "component_type": "Card",
        "quantity": 4,
        "supplier": str(supplier.pk),
        "warranty_type": "supplier",
        "warranty_months": 12,
    }, format="json")
    assert r.status_code == 201, r.content
    assert r.data["supplier_name"] == "Comp Supplier"
    w = Warranty.objects.get(component_id=r.data["id"])
    assert w.device == device and w.supplier == supplier
    assert w.warranty_type == "supplier" and w.months == 12
    # anchored at the device purchase (delivery) date
    assert str(w.start_date) == "2026-08-01" and str(w.end_date) == "2027-08-01"

    # component from another device is rejected on warranty create
    other = Device.objects.create(device_model=dm, asset_code="AST-COMP-2", serial_number="COMP-2")
    r2 = client.post("/api/warranties/", {
        "device": str(other.pk),
        "component": r.data["id"],
        "warranty_type": "supplier",
        "start_date": "2026-08-01",
        "end_date": "2027-08-01",
    }, format="json")
    assert r2.status_code == 400


# ── Device status machine (WF-05/07/08) ───────────────────────────────

from apps.accounts.models import AuditLog
from apps.assets.models import DeviceLifecycleEvent


@pytest.mark.django_db
def test_transition_writes_lifecycle_event_and_audit_log(admin_client, device):
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "in_production", "reason": "Assembly started"},
        format="json",
    )
    assert resp.status_code == 200, resp.content
    body = resp.json()
    assert body["status"] == "in_production"
    assert set(body["allowed_transitions"]) == {"in_stock", "rma"}

    event = DeviceLifecycleEvent.objects.get(
        device=device, event_type="status_change", from_value="procured"
    )
    assert event.to_value == "in_production"
    assert event.description == "Assembly started"
    assert event.performed_by.username == "labeladmin"

    log = AuditLog.objects.get(
        resource_type="device", resource_id=str(device.id), action="update"
    )
    assert log.user.username == "labeladmin"
    assert log.detail == {"from": "procured", "to": "in_production", "reason": "Assembly started"}


@pytest.mark.django_db
def test_invalid_transition_rejected_with_allowed_list(admin_client, device):
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "active", "reason": "skip ahead"},
        format="json",
    )
    assert resp.status_code == 400
    # error names the allowed next statuses
    assert "in_stock" in str(resp.json()["status"])
    # Only the registration journal entry exists — the rejected flip left none.
    assert not (
        DeviceLifecycleEvent.objects.filter(device=device)
        .exclude(description="Registered")
        .exists()
    )
    device.refresh_from_db()
    assert device.status == "procured"


@pytest.mark.django_db
def test_transition_requires_reason(admin_client, device):
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "in_stock"},
        format="json",
    )
    assert resp.status_code == 400
    resp = admin_client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "in_stock", "reason": "  "},
        format="json",
    )
    assert resp.status_code == 400


@pytest.mark.django_db
def test_transition_role_gate(device):
    tech = User.objects.create_user(username="statetech", password="x", role="technician")
    client = APIClient()
    client.force_authenticate(tech)
    resp = client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "in_stock", "reason": "QC passed"},
        format="json",
    )
    assert resp.status_code == 403

    warehouse = User.objects.create_user(username="statewh", password="x", role="warehouse")
    client.force_authenticate(warehouse)
    resp = client.post(
        f"/api/assets/devices/{device.id}/transition/",
        {"status": "in_stock", "reason": "QC passed"},
        format="json",
    )
    assert resp.status_code == 200, resp.content
    assert resp.json()["status"] == "in_stock"


@pytest.mark.django_db
def test_api_create_with_status_journals_initial_status(admin_client, device):
    """Registration may plant any status, but it must be journalled."""
    resp = admin_client.post("/api/assets/devices/", {
        "device_model": str(device.device_model_id),
        "serial_number": "SN-CREATE-ACTIVE",
        "status": "active",
    }, format="json")
    assert resp.status_code == 201, resp.content
    new_id = resp.json()["id"]
    assert resp.json()["status"] == "active"

    events = DeviceLifecycleEvent.objects.filter(device_id=new_id)
    assert events.count() == 1  # creation journals once, no double-fire
    event = events.get()
    assert event.event_type == "status_change"
    assert event.from_value == "" and event.to_value == "active"
    assert event.description == "Registered"

    log = AuditLog.objects.get(resource_type="device", resource_id=new_id)
    assert log.action == "create"
    assert log.detail == {"from": "", "to": "active", "reason": "Registered"}


@pytest.mark.django_db
def test_direct_create_journals_default_status(db):
    brand = Brand.objects.create(name="CreateBrand")
    dm = DeviceModel.objects.create(brand=brand, name="CR-1")
    dev = Device.objects.create(device_model=dm, serial_number="SN-CREATE-DEF")

    event = DeviceLifecycleEvent.objects.get(device=dev)
    assert event.from_value == "" and event.to_value == "procured"
    assert event.performed_by is None


@pytest.mark.django_db
def test_status_not_editable_via_plain_update(admin_client, device):
    resp = admin_client.patch(
        f"/api/assets/devices/{device.id}/", {"status": "active"}, format="json"
    )
    assert resp.status_code == 200, resp.content
    device.refresh_from_db()
    assert device.status == "procured"


@pytest.mark.django_db
def test_source_flag_default_and_filter(admin_client, device):
    detail = admin_client.get(f"/api/assets/devices/{device.id}/").json()
    assert detail["source"] == "third_party"

    resp = admin_client.post("/api/assets/devices/", {
        "device_model": str(device.device_model_id),
        "serial_number": "SN-INHOUSE-1",
        "source": "inhouse",
    }, format="json")
    assert resp.status_code == 201, resp.content
    assert resp.json()["source"] == "inhouse"

    rows = admin_client.get("/api/assets/devices/", {"source": "inhouse"}).json()
    results = rows.get("results", rows)
    assert [r["serial_number"] for r in results] == ["SN-INHOUSE-1"]


# ── Excel export (XC-01) ──────────────────────────────────────────────

import io as _io

from openpyxl import load_workbook as _load_workbook

from apps.accounts.models import AuditLog as _AuditLog

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _sheet_rows(resp):
    wb = _load_workbook(_io.BytesIO(resp.content), read_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


@pytest.mark.django_db
def test_devices_export_happy_path(admin_client, device):
    brand = device.device_model.brand
    dm2 = DeviceModel.objects.create(brand=brand, name="LB-77")
    Device.objects.create(device_model=dm2, serial_number="SN-EXPORT-2", status="in_stock")

    resp = admin_client.get("/api/assets/devices/export/")
    assert resp.status_code == 200, resp.content
    assert resp["Content-Type"] == XLSX_CT
    assert "assets-" in resp["Content-Disposition"] and ".xlsx" in resp["Content-Disposition"]

    rows = _sheet_rows(resp)
    assert rows[0][:3] == ["Asset Code", "Serial Number", "Name"]
    assert len(rows) == 3  # header + 2 devices

    log = _AuditLog.objects.filter(action="export", resource_type="device").latest("created_at")
    assert log.detail["count"] == 2
    assert log.user is not None


@pytest.mark.django_db
def test_devices_export_applies_filters(admin_client, device):
    brand = device.device_model.brand
    dm2 = DeviceModel.objects.create(brand=brand, name="LB-88")
    Device.objects.create(device_model=dm2, serial_number="SN-EXPORT-3", status="in_stock")

    resp = admin_client.get("/api/assets/devices/export/", {"status": "in_stock"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert len(rows) == 2  # header + 1 matching device
    assert rows[1][1] == "SN-EXPORT-3"

    log = _AuditLog.objects.filter(action="export", resource_type="device").latest("created_at")
    assert log.detail == {"count": 1, "params": {"status": "in_stock"}}


# ── Wave 5: ?flag= drill-downs + search parity ────────────────────────


@pytest.fixture
def flag_devices(device):
    dm = device.device_model  # `device` itself stays procured
    return {
        "active": Device.objects.create(device_model=dm, serial_number="SN-FLAG-ACT", status="active"),
        "installed": Device.objects.create(device_model=dm, serial_number="SN-FLAG-INST", status="installed"),
        "stock": Device.objects.create(device_model=dm, serial_number="SN-FLAG-STOCK", status="in_stock"),
    }


@pytest.mark.django_db
def test_device_flag_operational(admin_client, flag_devices):
    r = admin_client.get("/api/assets/devices/", {"flag": "operational", "page_size": 100})
    assert r.status_code == 200, r.content
    serials = {row["serial_number"] for row in r.data["results"]}
    assert serials == {"SN-FLAG-ACT", "SN-FLAG-INST"}


@pytest.mark.django_db
def test_device_flag_warranty_expired(admin_client, flag_devices):
    from datetime import timedelta

    from django.utils import timezone

    from apps.warranties.models import Warranty

    today = timezone.localdate()
    # Expired-only warranty → in the bucket.
    Warranty.objects.create(
        device=flag_devices["active"], warranty_type="client", status="expired",
        start_date=today - timedelta(days=400), end_date=today - timedelta(days=35), months=12,
    )
    # Expired AND active warranties → still covered, excluded.
    Warranty.objects.create(
        device=flag_devices["installed"], warranty_type="client", status="expired",
        start_date=today - timedelta(days=400), end_date=today - timedelta(days=35), months=12,
    )
    Warranty.objects.create(
        device=flag_devices["installed"], warranty_type="client", status="active",
        start_date=today, end_date=today + timedelta(days=365), months=12,
    )
    # flag_devices["stock"] has no warranties at all → excluded.

    r = admin_client.get("/api/assets/devices/", {"flag": "warranty_expired", "page_size": 100})
    assert r.status_code == 200, r.content
    serials = [row["serial_number"] for row in r.data["results"]]
    assert serials == ["SN-FLAG-ACT"]

    # Unknown flag values are ignored — full list comes back.
    r = admin_client.get("/api/assets/devices/", {"flag": "bogus", "page_size": 100})
    assert r.status_code == 200
    assert len(r.data["results"]) == 4


@pytest.mark.django_db
def test_device_flag_applies_to_export(admin_client, flag_devices):
    resp = admin_client.get("/api/assets/devices/export/", {"flag": "operational"})
    assert resp.status_code == 200, resp.content
    rows = _sheet_rows(resp)
    assert len(rows) == 3  # header + active + installed
    assert {r[1] for r in rows[1:]} == {"SN-FLAG-ACT", "SN-FLAG-INST"}

    log = _AuditLog.objects.filter(action="export", resource_type="device").latest("created_at")
    assert log.detail == {"count": 2, "params": {"flag": "operational"}}


@pytest.mark.django_db
def test_device_search_by_model_site_and_client(admin_client, device):
    from apps.clients.models import Client
    from apps.sites.models import Site

    site = Site.objects.create(name="Searchable Plaza", city="Lahore")
    client = Client.objects.create(name="Searchable Client Co")
    device.current_site = site
    device.assigned_client = client
    device.save(update_fields=["current_site", "assigned_client", "updated_at"])

    for term in ("LB-55", "Searchable Plaza", "Searchable Client"):
        r = admin_client.get("/api/assets/devices/", {"search": term})
        assert r.status_code == 200, r.content
        assert device.serial_number in {row["serial_number"] for row in r.data["results"]}, term

    r = admin_client.get("/api/assets/devices/", {"search": "no-such-thing-xyz"})
    assert r.data["results"] == []


@pytest.mark.django_db
def test_device_detail_exposes_project_contract(admin_client, device):
    from apps.teams.models import Project

    # No project — contract fields present but empty.
    r = admin_client.get(f"/api/assets/devices/{device.id}/")
    assert r.status_code == 200, r.content
    assert r.json()["project_contract_type"] is None
    assert r.json()["project_rental_end_date"] is None

    project = Project.objects.create(
        name="Rental Order", contract_type="rental", rental_end_date="2027-01-31"
    )
    device.project = project
    device.save(update_fields=["project", "updated_at"])

    r = admin_client.get(f"/api/assets/devices/{device.id}/")
    assert r.status_code == 200, r.content
    body = r.json()
    assert body["project_contract_type"] == "rental"
    assert body["project_rental_end_date"] == "2027-01-31"
