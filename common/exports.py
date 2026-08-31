"""Shared Excel export helpers (XC-01).

Every list-style export in the platform goes through :func:`xlsx_response`
so the workbooks look the same everywhere: bold frozen header row, sane
column widths and a dated filename. Views cap their querysets at
``EXPORT_MAX_ROWS`` and journal the export via :func:`log_export`.
"""

import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Hard cap on exported rows — exports are working documents, not backups.
EXPORT_MAX_ROWS = 5000


def _cell(value):
    """Coerce a value into something openpyxl can write.

    tz-aware datetimes are localised and stripped (Excel has no timezone
    concept); anything exotic (UUIDs, model instances) falls back to str().
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime.datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None, microsecond=0)
    if isinstance(value, (int, float, Decimal, str, datetime.date)):
        return value
    return str(value)


def xlsx_response(filename, sheet_name, columns, rows):
    """Build a styled .xlsx HttpResponse.

    ``columns`` is the list of header labels; ``rows`` an iterable of
    per-row value lists in the same order. ``filename`` is the base name
    (no extension) — today's date is appended automatically.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Export")[:31]

    ws.append([str(c) for c in columns])
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([_cell(v) for v in row])

    # Auto-ish widths from header length (clamped to stay readable).
    for idx, header in enumerate(columns, start=1):
        width = max(12, min(42, len(str(header)) + 8))
        ws.column_dimensions[get_column_letter(idx)].width = width

    dated_name = f"{filename}-{timezone.localdate().isoformat()}.xlsx"
    response = HttpResponse(content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{dated_name}"'
    wb.save(response)
    return response


def log_export(user, resource_type, count, params):
    """Journal an export in the audit trail (action=EXPORT)."""
    from apps.accounts.models import AuditLog

    AuditLog.objects.create(
        user=user if getattr(user, "is_authenticated", False) else None,
        action=AuditLog.Action.EXPORT,
        resource_type=resource_type,
        detail={"count": count, "params": params},
    )


def export_params(request):
    """The caller's query params as a plain JSON-safe dict for the audit row."""
    return {key: request.query_params.get(key) for key in request.query_params.keys()}
